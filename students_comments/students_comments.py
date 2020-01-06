# coding: UTF-8
import sys
import openpyxl
import random


def addtwodimdict(thedict, key_a, key_b, val):
    """ this is a function to add two dimension dict """
    if key_a in thedict:
        thedict[key_a].setdefault(key_b, []).append(val)
    else:
        thedict.update({key_a: {key_b: [val]}})
    return thedict


def process_comments_worksheet(comments_sheet, subject):
    comments_dict = {}
    for row in comments_sheet.iter_rows():
        if row[0].value == subject:
            addtwodimdict(comments_dict, row[1].value, row[2].value, row[3].value)
    return comments_dict


def process_results_worksheet(results_worksheet, comments_dict):
    comments_column = results_worksheet.max_column + 1
    for row in results_worksheet.iter_rows(min_row=5):
        score = row[len(row)-1].value
        if row[1].value == 'Male':
            if score.startswith("A"):
                results_worksheet.cell(row=row[0].row, column=comments_column).value = random.choice(comments_dict["He"]["A"]) + random.choice(comments_dict["He"]["D"])
            elif score.startswith(("B", "C")):
                results_worksheet.cell(row=row[0].row, column=comments_column).value = random.choice(comments_dict["He"]["B"]) + random.choice(comments_dict["He"]["D"])
            elif score.startswith(("D", "E", "U")):
                results_worksheet.cell(row=row[0].row, column=comments_column).value = random.choice(comments_dict["He"]["C"]) + random.choice(comments_dict["He"]["D"])
        elif row[1].value == 'Female':
            if score.startswith("A"):
                results_worksheet.cell(row=row[0].row, column=comments_column).value = random.choice(comments_dict["She"]["A"]) + random.choice(comments_dict["He"]["D"])
            elif score.startswith(("B", "C")):
                results_worksheet.cell(row=row[0].row, column=comments_column).value = random.choice(comments_dict["She"]["B"]) + random.choice(comments_dict["He"]["D"])
            elif score.startswith(("D", "E", "U")):
                results_worksheet.cell(row=row[0].row, column=comments_column).value = random.choice(comments_dict["She"]["C"]) + random.choice(comments_dict["He"]["D"])


# read worksheet
def main():
    # subject = "Chem"  # or "Bio"
    subject = sys.argv[1]
    # comments_file = "students_comments/comment_bank_for_chemANDbio.xlsx"
    comments_file = sys.argv[2]
    wb_comments = openpyxl.load_workbook(comments_file)
    comments_sheet = wb_comments.active
    comments_dict = process_comments_worksheet(comments_sheet, subject)

    # result_file = "students_comments/students.score.xlsx"
    result_file = sys.argv[3]
    results_workbook = openpyxl.load_workbook(result_file)
    results_worksheet = results_workbook.active
    process_results_worksheet(results_worksheet, comments_dict)
    results_workbook.save("example_copy.xlsx")


if __name__ == '__main__':
    main()